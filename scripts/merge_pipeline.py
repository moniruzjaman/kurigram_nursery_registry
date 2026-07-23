#!/usr/bin/env python3
"""
Reproducible Excel → JSON Merge Pipeline for Kurigram Nursery Registry

This script formalizes the one-off process documented in worklog.md into a
reproducible, version-controlled pipeline that can be re-run whenever the
source Excel files are updated.

Usage:
    python scripts/merge_pipeline.py [--cleaned EXCEL_PATH] [--registry EXCEL_PATH] [--output JSON_PATH]

Defaults:
    --cleaned   upload/Kurigram_Nursery_Cleaned_Data (2).xlsx
    --registry  upload/Kurigram_Plant_Nurseries_Expanded_Directory (1).xlsx
    --output    src/lib/nursery-data.json

The script:
  1. Reads the "Cleaned Data" sheet from the cleaned Excel file (pivot/inventory data)
  2. Reads the "nursery_registry" sheet from the expanded directory Excel file (registry data)
  3. Extracts mobile numbers from embedded text using regex
  4. Deduplicates nurseries from pivot data using owner+upazila as key
  5. Cross-references registry entries against pivot entries
  6. Merges the two data sources, filling gaps and preferring better values
  7. Validates the output against the schema in docs/DATA_SCHEMA.md
  8. Writes the merged JSON to the specified output path

Output JSON schema matches the Nursery interface in docs/DATA_SCHEMA.md:
  - registry_serial, owner, nursery_name_raw, address, mobile, mobile_source
  - upazila, latitude, longitude, gps_status, maps_link
  - fruit_seedlings, forest_seedlings, medicinal_seedlings, total_seedlings, main_variety
  - verification, priority
  - has_pivot_data, pivot_inventory, pivot_total_plants, pivot_total_seedlings, pivot_total_grafts
  - region, district
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_UPAZILAS = {
    unicodedata.normalize("NFC", "কুড়িগ্রাম সদর"),
    unicodedata.normalize("NFC", "উলিপুর"),
    unicodedata.normalize("NFC", "চর রাজিবপুর"),
    unicodedata.normalize("NFC", "চিলমারী"),
    unicodedata.normalize("NFC", "নাগেশ্বরী"),
    unicodedata.normalize("NFC", "ফুলবাড়ী"),
    unicodedata.normalize("NFC", "ভূরুঙ্গামারী"),
    unicodedata.normalize("NFC", "রাজারহাট"),
    unicodedata.normalize("NFC", "রৌমারী"),
}

MOBILE_REGEX = re.compile(r"(01[3-9]\d{8})")

# Kurigram District bounding box for GPS validation
LAT_MIN, LAT_MAX = 25.0, 26.0
LON_MIN, LON_MAX = 89.0, 90.0

# Default paths relative to project root
DEFAULT_CLEANED = "upload/Kurigram_Nursery_Cleaned_Data (2).xlsx"
DEFAULT_REGISTRY = "upload/Kurigram_Plant_Nurseries_Expanded_Directory (1).xlsx"
DEFAULT_OUTPUT = "src/lib/nursery-data.json"


# ---------------------------------------------------------------------------
# Step 1: Read pivot (cleaned) inventory data
# ---------------------------------------------------------------------------

def read_pivot_data(filepath: str) -> list[dict]:
    """
    Read the "Cleaned Data" sheet and return a list of raw row dicts.
    
    Header row is row 4 in the file:
      Col B (2): SL
      Col C (3): Nursery Name  (contains owner, address, mobile embedded)
      Col D (4): Region
      Col E (5): District
      Col F (6): Upazila
      Col G (7): Category
      Col H (8): Plant Name
      Col I (9): Age Group
      Col J (10): Total
      Col K (11): Seedlings
      Col L (12): Grafts
      Col M (13): Latitude
      Col N (14): Longitude
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Cleaned Data"]

    # Find header row dynamically (should be row 4 but we search)
    header_row = None
    header_map = {}
    for row_idx in range(1, 15):
        row_vals = {}
        for col_idx in range(1, 20):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                row_vals[col_idx] = str(v).strip()
        # Check for known header names
        known_headers = {"SL", "Nursery Name", "Region", "District", "Upazila",
                         "Category", "Plant Name", "Age Group", "Total",
                         "Seedlings", "Grafts", "Latitude", "Longitude"}
        matched = set(row_vals.values()) & known_headers
        if len(matched) >= 8:
            header_row = row_idx
            header_map = row_vals
            break

    if header_row is None:
        wb.close()
        raise ValueError(f"Could not find header row in {filepath}")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        sl = ws.cell(row=row_idx, column=2).value  # col B
        if sl is None:
            continue  # skip empty rows

        row_dict = {
            "sl": sl,
            "nursery_name_raw": ws.cell(row=row_idx, column=3).value or "",
            "region": ws.cell(row=row_idx, column=4).value or "",
            "district": ws.cell(row=row_idx, column=5).value or "",
            "upazila": ws.cell(row=row_idx, column=6).value or "",
            "category": ws.cell(row=row_idx, column=7).value or "",
            "plant_name": ws.cell(row=row_idx, column=8).value,
            "age_group": ws.cell(row=row_idx, column=9).value or "",
            "total": ws.cell(row=row_idx, column=10).value or 0,
            "seedlings": ws.cell(row=row_idx, column=11).value or 0,
            "grafts": ws.cell(row=row_idx, column=12).value or 0,
            "latitude": ws.cell(row=row_idx, column=13).value,
            "longitude": ws.cell(row=row_idx, column=14).value,
        }
        rows.append(row_dict)

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Step 2: Read registry data
# ---------------------------------------------------------------------------

def read_registry_data(filepath: str) -> list[dict]:
    """
    Read the "nursery_registry" sheet and return a list of raw row dicts.
    
    Header row is row 12 in the file:
      Col A (1): ক্র. (serial number)
      Col B (2): উপজেলা (upazila)
      Col C (3): নার্সারী/মালিকের নাম (nursery/owner name)
      Col D (4): ঠিকানা / গ্রাম (address/village)
      Col E (5): মোবাইল নম্বর (mobile number)
      Col F (6): যাচাই (verification)
      Col G (7): অক্ষাংশ (latitude)
      Col H (8): দ্রাঘিমাংশ (longitude)
      Col I (9): GPS স্ট্যাটাস (GPS status)
      Col J (10): মূল GPS রেকর্ড (original GPS record)
      Col K (11): Google Maps লিংক (maps link)
      Col L (12): ফলদ চারা (fruit seedlings)
      Col M (13): বনজ চারা (forest seedlings)
      Col N (14): ঔষধি চারা (medicinal seedlings)
      Col O (15): মোট চারা (total seedlings)
      Col P (16): প্রধান জাত (main variety)
      Col Q (17): নিবন্ধন নং (registration number)
      Col R (18): নিবন্ধন বছর (registration year)
      Col S (19): জমি (শতক) (land area)
      Col T (20): সর্বশেষ যাচাই (last verification)
      Col U (21): ডেটা সংশোধন নোট (correction note)
      Col V (22): অগ্রাধিকার (priority)
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb["nursery_registry"]

    # Find header row dynamically (should be row 12 but we search for 'ক্র.')
    header_row = None
    for row_idx in range(1, 20):
        val_a = ws.cell(row=row_idx, column=1).value
        val_b = ws.cell(row=row_idx, column=2).value
        if val_a == "ক্র." and val_b == "উপজেলা":
            header_row = row_idx
            break

    if header_row is None:
        wb.close()
        raise ValueError(f"Could not find header row in {filepath}")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        serial = ws.cell(row=row_idx, column=1).value
        if serial is None:
            continue  # skip empty rows
        # Skip summary/total rows (non-numeric serial values like 'সর্বমোট')
        try:
            serial_num = int(float(serial))
        except (ValueError, TypeError):
            continue

        row_dict = {
            "serial": serial_num,
            "upazila": ws.cell(row=row_idx, column=2).value or "",
            "nursery_name_raw": ws.cell(row=row_idx, column=3).value or "",
            "address": ws.cell(row=row_idx, column=4).value,
            "mobile": ws.cell(row=row_idx, column=5).value,
            "verification": ws.cell(row=row_idx, column=6).value,
            "latitude": ws.cell(row=row_idx, column=7).value,
            "longitude": ws.cell(row=row_idx, column=8).value,
            "gps_status": ws.cell(row=row_idx, column=9).value,
            "original_gps": ws.cell(row=row_idx, column=10).value,
            "maps_link": ws.cell(row=row_idx, column=11).value,
            "fruit_seedlings": ws.cell(row=row_idx, column=12).value or 0,
            "forest_seedlings": ws.cell(row=row_idx, column=13).value or 0,
            "medicinal_seedlings": ws.cell(row=row_idx, column=14).value or 0,
            "total_seedlings": ws.cell(row=row_idx, column=15).value or 0,
            "main_variety": ws.cell(row=row_idx, column=16).value,
            "priority": ws.cell(row=row_idx, column=22).value,
        }
        # Normalize numeric fields
        for key in ("fruit_seedlings", "forest_seedlings", "medicinal_seedlings", "total_seedlings"):
            if row_dict[key] is not None:
                row_dict[key] = int(float(row_dict[key]))
            else:
                row_dict[key] = 0
        if row_dict["latitude"] is not None:
            row_dict["latitude"] = float(row_dict["latitude"])
        if row_dict["longitude"] is not None:
            row_dict["longitude"] = float(row_dict["longitude"])
        # Normalize mobile to string or null
        if row_dict["mobile"] is not None:
            mobile_str = str(row_dict["mobile"]).strip()
            match = MOBILE_REGEX.search(mobile_str)
            if match:
                row_dict["mobile"] = match.group(1)
            else:
                row_dict["mobile"] = None
        else:
            row_dict["mobile"] = None

        rows.append(row_dict)

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Helper: normalize text for fuzzy matching
# ---------------------------------------------------------------------------

def normalize_text(text: str) -> str:
    """
    Normalize Bengali text for matching and validation.
    
    Handles two critical issues with Bengali Unicode:
    1. Whitespace: collapse multiple spaces/newlines to single space
    2. Unicode normalization: Bengali has precomposed and decomposed forms
       (e.g., ড় can be U+09DC or U+09A1+U+09BC). NFC normalization ensures
       consistent representation for comparison.
    
    All text comparisons (upazila validation, name matching, etc.) should
    use this function.
    """
    if not text:
        return ""
    # First: NFC Unicode normalization (precomposed form for consistency)
    text = unicodedata.normalize("NFC", text)
    # Then: collapse multiple spaces/newlines to single space
    return re.sub(r"\s+", " ", text).strip()


def extract_name_parts(raw_name: str) -> list[str]:
    """
    Extract candidate name parts from a raw nursery name string.
    
    The raw name can be in various formats:
    - "অমল চন্দ্র, দূর্গাপুর" → ["অমল চন্দ্র"]
    - "লতা নার্সারী মোঃ আঃ জলিল" → ["লতা নার্সারী মোঃ আঃ জলিল", "মোঃ আঃ জলিল"]
    - "গ্রিণ বাংলা নার্সারী,মোঃ আবুল কাশেম, শিবরাম" → ["গ্রিণ বাংলা নার্সারী", "মোঃ আবুল কাশেম"]
    - "খোকা নার্সারী মোঃ রহিম উদ্দিন, আধাগ্রাম" → ["খোকা নার্সারী মোঃ রহিম উদ্দিন", "মোঃ রহিম উদ্দিন"]
    
    Returns list of normalized candidate strings.
    """
    # Split by commas
    parts = [normalize_text(p) for p in raw_name.split(",") if normalize_text(p)]
    candidates = []
    
    for part in parts:
        # Strip mobile numbers from the part
        clean = MOBILE_REGEX.sub("", part).strip()
        if clean:
            candidates.append(normalize_text(clean))
        # Also check for person-name patterns within the part
        # মোঃ (Md.), শ্রী ( Sri), মোসাঃ (Mosammat) prefixes indicate person names
        person_match = re.search(r"(মোঃ[\s\.]*[^,]+|শ্রী[\s\.]*[^,]+|মোসাঃ[\s\.]*[^,]+|মোঃ\s*আঃ[^,]+)", part)
        if person_match:
            person_name = normalize_text(person_match.group(1))
            if person_name and person_name not in candidates:
                candidates.append(person_name)
    
    # Also add the full normalized name (without mobile) as a candidate
    full_clean = normalize_text(MOBILE_REGEX.sub("", raw_name).strip())
    if full_clean and full_clean not in candidates:
        candidates.append(full_clean)
    
    return candidates


# ---------------------------------------------------------------------------
# Step 3: Extract unique nurseries from pivot data
# ---------------------------------------------------------------------------

def dedup_pivot_nurseries(pivot_rows: list[dict]) -> dict[str, dict]:
    """
    Group pivot rows by owner+upazila dedup key, extracting:
    - owner name (from nursery_name_raw, first comma-separated part, cleaned)
    - mobile (extracted from embedded text via regex)
    - latitude/longitude (from the first row that has valid values)
    - inventory items grouped under each nursery
    - name_parts: all candidate name strings for cross-referencing
    
    Returns dict keyed by (owner, upazila) tuple string.
    """
    nurseries = {}

    for row in pivot_rows:
        raw_name = row["nursery_name_raw"]
        # Extract owner: first part before comma, cleaned (no mobile)
        parts = raw_name.split(",")
        owner_raw = parts[0].strip() if parts else raw_name.strip()
        # Remove embedded mobile from owner
        owner = normalize_text(MOBILE_REGEX.sub("", owner_raw).strip())

        # Extract mobile from embedded text
        mobile_match = MOBILE_REGEX.search(raw_name)
        extracted_mobile = mobile_match.group(1) if mobile_match else None

        upazila = normalize_text(row["upazila"])
        key = f"{owner}||{upazila}"

        inventory_item = {
            "category": row["category"],
            "plant_name": row["plant_name"],
            "age_group": row["age_group"],
            "total": int(float(row["total"])) if row["total"] else 0,
            "seedlings": int(float(row["seedlings"])) if row["seedlings"] else 0,
            "grafts": int(float(row["grafts"])) if row["grafts"] else 0,
        }

        if key not in nurseries:
            nurseries[key] = {
                "owner": owner,
                "upazila": upazila,
                "nursery_name_raw": normalize_text(raw_name),
                "name_parts": extract_name_parts(raw_name),
                "region": normalize_text(row["region"]),
                "district": normalize_text(row["district"]),
                "mobile": extracted_mobile,
                "mobile_source": "extracted" if extracted_mobile else None,
                "latitude": row["latitude"],
                "longitude": row["longitude"],
                "pivot_inventory": [],
                "pivot_total_plants": 0,
                "pivot_total_seedlings": 0,
                "pivot_total_grafts": 0,
            }
        else:
            # Prefer better mobile/GPS from later rows
            entry = nurseries[key]
            if entry["mobile"] is None and extracted_mobile:
                entry["mobile"] = extracted_mobile
                entry["mobile_source"] = "extracted"
            if entry["latitude"] is None and row["latitude"] is not None:
                entry["latitude"] = row["latitude"]
            if entry["longitude"] is None and row["longitude"] is not None:
                entry["longitude"] = row["longitude"]

        nurseries[key]["pivot_inventory"].append(inventory_item)
        nurseries[key]["pivot_total_plants"] += inventory_item["total"]
        nurseries[key]["pivot_total_seedlings"] += inventory_item["seedlings"]
        nurseries[key]["pivot_total_grafts"] += inventory_item["grafts"]

    return nurseries


# ---------------------------------------------------------------------------
# Step 4: Cross-reference and merge
# ---------------------------------------------------------------------------

def find_best_match(pivot_entry: dict, registry_lookup: dict, 
                    matched_registry_indices: set) -> tuple[str, dict] | tuple[None, None]:
    """
    Find the best matching registry entry for a pivot entry using multiple strategies:
    
    1. Exact match: pivot owner == registry first-part + same upazila
    2. Part match: any pivot name_part == any registry name_part + same upazila
    3. Substring match: pivot name_part contained in registry name or vice versa + same upazila
    
    Returns (match_key, registry_entry) or (None, None).
    """
    pivot_parts = pivot_entry.get("name_parts", [pivot_entry["owner"]])
    pivot_upazila = pivot_entry["upazila"]
    
    # Strategy 1: Exact key match (owner||upazila)
    for part in pivot_parts:
        key = f"{part}||{pivot_upazila}"
        if key in registry_lookup and key not in matched_registry_indices:
            return key, registry_lookup[key]
    
    # Strategy 2: Match any pivot part against any registry part (same upazila)
    for reg_key, reg_entry in registry_lookup.items():
        if reg_key in matched_registry_indices:
            continue
        if reg_entry["upazila"] != pivot_upazila:
            continue
        
        reg_parts = extract_name_parts(reg_entry["nursery_name_raw"])
        
        # Check if any pivot part matches any registry part
        for pp in pivot_parts:
            for rp in reg_parts:
                if pp == rp:
                    return reg_key, reg_entry
    
    # Strategy 3: Substring match (one contains the other)
    for reg_key, reg_entry in registry_lookup.items():
        if reg_key in matched_registry_indices:
            continue
        if reg_entry["upazila"] != pivot_upazila:
            continue
        
        reg_parts = extract_name_parts(reg_entry["nursery_name_raw"])
        
        for pp in pivot_parts:
            for rp in reg_parts:
                # Check if one name contains the other (for partial matches)
                if len(pp) > 3 and len(rp) > 3:
                    if pp in rp or rp in pp:
                        return reg_key, reg_entry
    
    return None, None


def merge_data(pivot_nurseries: dict, registry_rows: list[dict]) -> list[dict]:
    """
    Cross-reference registry entries with pivot nurseries using multi-strategy matching.
    
    Strategy:
    1. Match registry entries to pivot entries using owner name + upazila (multi-strategy)
    2. For matched entries: combine data, preferring better values
    3. For unmatched pivot entries: add as inventory-only entries (no registry_serial)
    4. For unmatched registry entries: add with registry data but no pivot data
    
    Merge rules:
    - Mobile: prefer registry mobile over extracted mobile
    - GPS: prefer pivot GPS if registry has default/missing coordinates
    - All registry fields are kept; pivot fills gaps
    """
    # Build registry lookup: key = "first_part||upazila", with full entry
    registry_lookup = {}
    for reg in registry_rows:
        raw_name = normalize_text(reg["nursery_name_raw"])
        parts = raw_name.split(",")
        first_part = normalize_text(parts[0]) if parts else raw_name
        # Clean mobile from first part
        first_part_clean = normalize_text(MOBILE_REGEX.sub("", first_part).strip())
        upazila = normalize_text(reg["upazila"])
        key = f"{first_part_clean}||{upazila}"
        
        # Store with normalized fields for matching
        reg_normalized = dict(reg)
        reg_normalized["nursery_name_raw"] = raw_name
        reg_normalized["upazila"] = upazila
        
        # If key collision, keep the first entry (shouldn't happen in well-structured data)
        if key not in registry_lookup:
            registry_lookup[key] = reg_normalized
        # Also add alternative keys from other name parts
        reg_parts = extract_name_parts(raw_name)
        for alt_part in reg_parts:
            alt_key = f"{alt_part}||{upazila}"
            if alt_key not in registry_lookup:
                registry_lookup[alt_key] = reg_normalized

    # Track matched registry entries by their primary key
    matched_registry_indices = set()

    merged = []

    # Step 1: Process pivot entries (our primary data source)
    for pivot_key, pivot_entry in pivot_nurseries.items():
        # Try multi-strategy matching
        match_key, registry_entry = find_best_match(
            pivot_entry, registry_lookup, matched_registry_indices
        )

        if registry_entry:
            matched_registry_indices.add(match_key)
            # Also mark the primary key and all alternative keys
            first_part = normalize_text(pivot_entry["owner"])
            primary_key = f"{first_part}||{pivot_entry['upazila']}"
            matched_registry_indices.add(primary_key)
            entry = build_merged_entry(registry_entry, pivot_entry, matched=True)
        else:
            # Pivot-only entry (not in registry)
            entry = build_pivot_only_entry(pivot_entry)

        merged.append(entry)

    # Step 2: Add unmatched registry entries (no pivot inventory)
    # Use a set of actual registry entries (not just keys) to avoid duplicates
    matched_reg_entries = set()
    for key in matched_registry_indices:
        if key in registry_lookup:
            matched_reg_entries.add(registry_lookup[key]["serial"])

    for reg_key, reg_entry in registry_lookup.items():
        if reg_entry["serial"] not in matched_reg_entries:
            # Only add each unique registry entry once
            matched_reg_entries.add(reg_entry["serial"])
            entry = build_registry_only_entry(reg_entry)
            merged.append(entry)

    # Sort by registry_serial (nulls last), then by owner
    merged.sort(key=lambda e: (e["registry_serial"] is None, e["registry_serial"] or 0, e["owner"]))

    return merged


def build_merged_entry(registry_entry: dict, pivot_entry: dict, matched: bool = True) -> dict:
    """Build a merged nursery entry from registry and pivot data."""
    
    # Determine best mobile
    mobile = registry_entry.get("mobile")
    mobile_source = "registry" if mobile else None
    
    if mobile is None and pivot_entry.get("mobile"):
        mobile = pivot_entry["mobile"]
        mobile_source = "extracted"
    
    # Determine best GPS
    lat = registry_entry.get("latitude")
    lon = registry_entry.get("longitude")
    
    # If registry GPS is default/missing, try pivot
    if (lat is None or lon is None) or (lat == 25.560833 and lon == 89.831111):
        # Default coordinates for Kurigram Sadar — use pivot if available and different
        pivot_lat = pivot_entry.get("latitude")
        pivot_lon = pivot_entry.get("longitude")
        if pivot_lat is not None and pivot_lon is not None:
            if pivot_lat != lat or pivot_lon != lon:
                lat = pivot_lat
                lon = pivot_lon

    # Build address
    address = registry_entry.get("address")
    if address is None:
        # Try to extract from pivot raw name (parts after first comma)
        parts = pivot_entry["nursery_name_raw"].split(",")
        if len(parts) > 1:
            address = parts[1].strip()

    entry = {
        "registry_serial": int(registry_entry["serial"]) if registry_entry.get("serial") else None,
        "owner": pivot_entry.get("owner", registry_entry.get("nursery_name_raw", "").split(",")[0].strip()),
        "nursery_name_raw": registry_entry.get("nursery_name_raw", pivot_entry.get("nursery_name_raw", "")),
        "address": address,
        "mobile": mobile,
        "mobile_source": mobile_source,
        "upazila": registry_entry.get("upazila", pivot_entry.get("upazila", "")),
        "latitude": lat,
        "longitude": lon,
        "gps_status": registry_entry.get("gps_status"),
        "maps_link": registry_entry.get("maps_link"),
        "fruit_seedlings": registry_entry.get("fruit_seedlings", 0),
        "forest_seedlings": registry_entry.get("forest_seedlings", 0),
        "medicinal_seedlings": registry_entry.get("medicinal_seedlings", 0),
        "total_seedlings": registry_entry.get("total_seedlings", 0),
        "main_variety": registry_entry.get("main_variety"),
        "verification": registry_entry.get("verification"),
        "priority": registry_entry.get("priority"),
        "has_pivot_data": True,
        "pivot_inventory": pivot_entry.get("pivot_inventory", []),
        "pivot_total_plants": pivot_entry.get("pivot_total_plants", 0),
        "pivot_total_seedlings": pivot_entry.get("pivot_total_seedlings", 0),
        "pivot_total_grafts": pivot_entry.get("pivot_total_grafts", 0),
        "region": pivot_entry.get("region", "রংপুর"),
        "district": pivot_entry.get("district", "কুড়িগ্রাম"),
    }

    return entry


def build_pivot_only_entry(pivot_entry: dict) -> dict:
    """Build entry for nurseries only in pivot data (not in registry)."""
    return {
        "registry_serial": None,
        "owner": pivot_entry["owner"],
        "nursery_name_raw": pivot_entry["nursery_name_raw"],
        "address": None,
        "mobile": pivot_entry.get("mobile"),
        "mobile_source": pivot_entry.get("mobile_source"),
        "upazila": pivot_entry["upazila"],
        "latitude": pivot_entry.get("latitude"),
        "longitude": pivot_entry.get("longitude"),
        "gps_status": None,
        "maps_link": None,
        "fruit_seedlings": 0,
        "forest_seedlings": 0,
        "medicinal_seedlings": 0,
        "total_seedlings": 0,
        "main_variety": None,
        "verification": None,
        "priority": None,
        "has_pivot_data": True,
        "pivot_inventory": pivot_entry.get("pivot_inventory", []),
        "pivot_total_plants": pivot_entry.get("pivot_total_plants", 0),
        "pivot_total_seedlings": pivot_entry.get("pivot_total_seedlings", 0),
        "pivot_total_grafts": pivot_entry.get("pivot_total_grafts", 0),
        "region": pivot_entry.get("region", "রংপুর"),
        "district": pivot_entry.get("district", "কুড়িগ্রাম"),
    }


def build_registry_only_entry(registry_entry: dict) -> dict:
    """Build entry for nurseries only in registry (no pivot inventory)."""
    raw_name = registry_entry.get("nursery_name_raw", "")
    parts = raw_name.split(",")
    owner = parts[0].strip() if parts else raw_name.strip()

    return {
        "registry_serial": int(registry_entry["serial"]) if registry_entry.get("serial") else None,
        "owner": owner,
        "nursery_name_raw": raw_name,
        "address": registry_entry.get("address"),
        "mobile": registry_entry.get("mobile"),
        "mobile_source": "registry" if registry_entry.get("mobile") else None,
        "upazila": registry_entry.get("upazila", ""),
        "latitude": registry_entry.get("latitude"),
        "longitude": registry_entry.get("longitude"),
        "gps_status": registry_entry.get("gps_status"),
        "maps_link": registry_entry.get("maps_link"),
        "fruit_seedlings": registry_entry.get("fruit_seedlings", 0),
        "forest_seedlings": registry_entry.get("forest_seedlings", 0),
        "medicinal_seedlings": registry_entry.get("medicinal_seedlings", 0),
        "total_seedlings": registry_entry.get("total_seedlings", 0),
        "main_variety": registry_entry.get("main_variety"),
        "verification": registry_entry.get("verification"),
        "priority": registry_entry.get("priority"),
        "has_pivot_data": False,
        "pivot_inventory": [],
        "pivot_total_plants": 0,
        "pivot_total_seedlings": 0,
        "pivot_total_grafts": 0,
        "region": "রংপুর",
        "district": "কুড়িগ্রাম",
    }


# ---------------------------------------------------------------------------
# Step 5: Validate output against schema
# ---------------------------------------------------------------------------

def validate_merged_data(data: list[dict]) -> list[str]:
    """
    Validate merged data against the schema rules in DATA_SCHEMA.md.
    Returns a list of warning/error messages (empty if all valid).
    """
    issues = []

    for entry in data:
        idx = entry.get("registry_serial") or "?"
        owner = entry.get("owner", "?")

        # owner must be non-empty
        if not entry.get("owner"):
            issues.append(f"[serial={idx}] owner is empty")

        # upazila must be one of 9 valid values (normalize for comparison)
        upazila_normalized = normalize_text(entry.get("upazila", ""))
        if upazila_normalized not in VALID_UPAZILAS:
            issues.append(f"[serial={idx}, owner={owner}] invalid upazila: {upazila_normalized}")

        # region must be রংপুর (normalize for comparison)
        region_normalized = normalize_text(entry.get("region", ""))
        if region_normalized != "রংপুর":
            issues.append(f"[serial={idx}] region should be রংপুর: {region_normalized}")

        # district must be কুড়িগ্রাম (normalize for comparison)
        district_normalized = normalize_text(entry.get("district", ""))
        if district_normalized != "কুড়িগ্রাম":
            issues.append(f"[serial={idx}] district should be কুড়িগ্রাম: {district_normalized}")

        # mobile validation
        if entry.get("mobile") is not None:
            if not MOBILE_REGEX.fullmatch(entry["mobile"]):
                issues.append(f"[serial={idx}] invalid mobile: {entry['mobile']}")

        # GPS bounding box validation
        lat = entry.get("latitude")
        lon = entry.get("longitude")
        if lat is not None and lon is not None and lat != 0 and lon != 0:
            if not (LAT_MIN <= lat <= LAT_MAX):
                issues.append(f"[serial={idx}] latitude out of range: {lat}")
            if not (LON_MIN <= lon <= LON_MAX):
                issues.append(f"[serial={idx}] longitude out of range: {lon}")

        # Inventory item validation
        for inv in entry.get("pivot_inventory", []):
            if inv["total"] != inv["seedlings"] + inv["grafts"]:
                issues.append(
                    f"[serial={idx}] inventory mismatch: "
                    f"total={inv['total']} != seedlings({inv['seedlings']})+grafts({inv['grafts']})"
                )

        # Pivot totals validation
        if entry.get("has_pivot_data") and entry.get("pivot_inventory"):
            sum_total = sum(inv["total"] for inv in entry["pivot_inventory"])
            sum_seedlings = sum(inv["seedlings"] for inv in entry["pivot_inventory"])
            sum_grafts = sum(inv["grafts"] for inv in entry["pivot_inventory"])

            if entry["pivot_total_plants"] != sum_total:
                issues.append(
                    f"[serial={idx}] pivot_total_plants={entry['pivot_total_plants']} "
                    f"!= sum(total)={sum_total}"
                )
            if entry["pivot_total_seedlings"] != sum_seedlings:
                issues.append(
                    f"[serial={idx}] pivot_total_seedlings={entry['pivot_total_seedlings']} "
                    f"!= sum(seedlings)={sum_seedlings}"
                )
            if entry["pivot_total_grafts"] != sum_grafts:
                issues.append(
                    f"[serial={idx}] pivot_total_grafts={entry['pivot_total_grafts']} "
                    f"!= sum(grafts)={sum_grafts}"
                )

    return issues


# ---------------------------------------------------------------------------
# Step 6: Write output
# ---------------------------------------------------------------------------

def write_json(data: list[dict], output_path: str) -> None:
    """Write merged data to JSON file."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Written {len(data)} nursery entries to {output_path}")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline(cleaned_path: str, registry_path: str, output_path: str) -> None:
    """Execute the full Excel → JSON merge pipeline."""
    
    print("=" * 60)
    print("Kurigram Nursery Registry — Excel → JSON Merge Pipeline")
    print("=" * 60)

    # Step 1: Read pivot data
    print("\n[Step 1] Reading pivot (cleaned) inventory data...")
    pivot_rows = read_pivot_data(cleaned_path)
    print(f"  -> {len(pivot_rows)} inventory rows read")

    # Step 2: Read registry data
    print("\n[Step 2] Reading registry data...")
    registry_rows = read_registry_data(registry_path)
    print(f"  -> {len(registry_rows)} registry entries read")

    # Step 3: Deduplicate pivot data into unique nurseries
    print("\n[Step 3] Deduplicating pivot data into unique nurseries...")
    pivot_nurseries = dedup_pivot_nurseries(pivot_rows)
    print(f"  -> {len(pivot_nurseries)} unique nurseries from pivot data")
    
    mobile_count = sum(1 for n in pivot_nurseries.values() if n.get("mobile"))
    print(f"  -> {mobile_count} nurseries with mobile numbers (extracted)")

    # Step 4: Cross-reference and merge
    print("\n[Step 4] Cross-referencing and merging...")
    merged = merge_data(pivot_nurseries, registry_rows)
    
    matched = sum(1 for e in merged if e.get("has_pivot_data") and e.get("registry_serial") is not None)
    pivot_only = sum(1 for e in merged if e.get("has_pivot_data") and e.get("registry_serial") is None)
    registry_only = sum(1 for e in merged if not e.get("has_pivot_data"))
    
    print(f"  -> {matched} matched (registry + pivot)")
    print(f"  -> {pivot_only} pivot-only (no registry entry)")
    print(f"  -> {registry_only} registry-only (no pivot inventory)")
    print(f"  -> {len(merged)} total merged entries")

    total_mobile = sum(1 for e in merged if e.get("mobile"))
    valid_gps = sum(1 for e in merged if e.get("latitude") and e.get("longitude")
                    and LAT_MIN <= e["latitude"] <= LAT_MAX
                    and LON_MIN <= e["longitude"] <= LON_MAX)
    total_plants = sum(e.get("pivot_total_plants", 0) for e in merged)
    total_seedlings = sum(e.get("pivot_total_seedlings", 0) for e in merged)
    total_grafts = sum(e.get("pivot_total_grafts", 0) for e in merged)
    
    print(f"  -> {total_mobile} nurseries with mobile numbers")
    print(f"  -> {valid_gps} nurseries with valid GPS")
    print(f"  -> {total_plants:,} total plants")
    print(f"  -> {total_seedlings:,} total seedlings")
    print(f"  -> {total_grafts:,} total grafts")

    # Step 5: Validate
    print("\n[Step 5] Validating merged data...")
    issues = validate_merged_data(merged)
    if issues:
        print(f"  WARNING: {len(issues)} validation issues found:")
        for issue in issues[:20]:
            print(f"    - {issue}")
        if len(issues) > 20:
            print(f"    ... and {len(issues) - 20} more issues")
    else:
        print("  OK: All validation checks passed")

    # Step 6: Write output
    print("\n[Step 6] Writing output JSON...")
    write_json(merged, output_path)

    # Also write a copy to upload/merged_nursery_data.json for archival
    project_root = os.path.dirname(os.path.dirname(output_path))
    upload_path = os.path.join(project_root, "upload", "merged_nursery_data.json")
    if os.path.isdir(os.path.dirname(upload_path)):
        write_json(merged, upload_path)

    print("\n" + "=" * 60)
    print("Pipeline complete!")
    print("=" * 60)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Reproducible Excel -> JSON Merge Pipeline for Kurigram Nursery Registry"
    )
    parser.add_argument(
        "--cleaned", default=DEFAULT_CLEANED,
        help="Path to the cleaned nursery data Excel file (default: %(default)s)"
    )
    parser.add_argument(
        "--registry", default=DEFAULT_REGISTRY,
        help="Path to the expanded directory Excel file (default: %(default)s)"
    )
    parser.add_argument(
        "--output", default=DEFAULT_OUTPUT,
        help="Path for the output JSON file (default: %(default)s)"
    )
    parser.add_argument(
        "--project-root", default=None,
        help="Project root directory (auto-detected if not specified)"
    )

    args = parser.parse_args()

    # Auto-detect project root (directory containing this script's parent)
    if args.project_root is None:
        script_dir = Path(__file__).resolve().parent
        args.project_root = str(script_dir.parent)

    # Resolve paths relative to project root
    cleaned_path = os.path.join(args.project_root, args.cleaned)
    registry_path = os.path.join(args.project_root, args.registry)
    output_path = os.path.join(args.project_root, args.output)

    # Verify input files exist
    if not os.path.isfile(cleaned_path):
        print(f"ERROR: Cleaned data file not found: {cleaned_path}")
        sys.exit(1)
    if not os.path.isfile(registry_path):
        print(f"ERROR: Registry data file not found: {registry_path}")
        sys.exit(1)

    run_pipeline(cleaned_path, registry_path, output_path)


if __name__ == "__main__":
    main()
